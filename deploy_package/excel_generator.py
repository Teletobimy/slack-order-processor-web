# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
import os

class ExcelGenerator:
    def __init__(self, config_path: str = "config.json"):
        """Excel 생성 클래스 초기화"""
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)
        
        # Template 구조 로드
        self.template_structure = self.load_template_structure()
        
    def load_template_structure(self) -> Dict[str, Any]:
        """Template 구조 로드"""
        try:
            with open(self.config['template'], 'r', encoding='utf-8') as f:
                template = json.load(f)
            return template
        except Exception as e:
            print(f"Template 로드 오류: {e}")
            return {}
    
    def create_excel_files_by_brand(self, aggregated_data: Dict[str, Any], output_dir: str) -> List[str]:
        """
        브랜드별로 Excel 파일 생성
        """
        created_files = []
        aggregated_by_brand = aggregated_data.get("aggregated_by_brand", {})
        
        if not aggregated_by_brand:
            print("집계된 브랜드 데이터가 없습니다.")
            return created_files
        
        try:
            for brand_name, products in aggregated_by_brand.items():
                if not products:  # 제품이 없는 브랜드는 건너뛰기
                    continue
                
                # 브랜드별 파일명 생성
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{brand_name}_주문서_{timestamp}.xlsx"
                filepath = os.path.join(output_dir, filename)
                
                # 새 워크북 생성
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "주문서입력"
                
                # 헤더 설정
                self.setup_headers(ws)
                
                # 데이터 입력
                self.fill_data(ws, products)
                
                # 스타일 적용
                self.apply_styles(ws, len(products))
                
                # 요약 시트 추가
                self.create_summary_sheet(wb, aggregated_data, brand_name)
                
                # 파일 저장
                wb.save(filepath)
                created_files.append(filepath)
                print(f"Excel 파일 생성 완료: {filename}")
            
            return created_files
            
        except Exception as e:
            print(f"Excel 생성 오류: {e}")
            return created_files
    
    def setup_headers(self, ws):
        """헤더 설정"""
        headers = {
            'A1': '일자',
            'B1': '순번', 
            'C1': '거래처코드',
            'D1': '거래처명',
            'E1': '출하창고',
            'F1': '담당자',
            'G1': '품목코드',
            'H1': '품목명',
            'I1': '규격',
            'J1': '수량',
            'K1': '사용유형',
            'L1': '적요'
        }
        
        for cell, value in headers.items():
            ws[cell] = value
    
    def fill_data(self, ws, products: List[Dict[str, Any]]):
        """데이터 입력"""
        today = datetime.now().strftime('%Y-%m-%d')
        warehouse_code = self.config.get('warehouse_code', '100')
        
        for i, product in enumerate(products, start=2):
            # 필수 컬럼 채우기
            ws[f'A{i}'] = today  # 일자
            ws[f'B{i}'] = i - 1  # 순번 (1부터 시작)
            ws[f'E{i}'] = warehouse_code  # 출하창고
            ws[f'G{i}'] = product['품목코드']  # 품목코드
            ws[f'J{i}'] = product['총_수량']  # 수량
            ws[f'L{i}'] = product.get('적요', '출고 처리')  # 적요
            
            # 빈 컬럼 (C, D, F, H, I, K)
            ws[f'C{i}'] = ''  # 거래처코드
            ws[f'D{i}'] = ''  # 거래처명
            ws[f'F{i}'] = ''  # 담당자
            ws[f'H{i}'] = ''  # 품목명
            ws[f'I{i}'] = ''  # 규격
            ws[f'K{i}'] = ''  # 사용유형
    
    def apply_styles(self, ws, data_rows: int):
        """스타일 적용"""
        # 헤더 스타일
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더에 스타일 적용
        for col in range(1, 13):  # A부터 L까지
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
        
        # 데이터 영역 스타일
        data_alignment = Alignment(horizontal='center', vertical='center')
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 데이터 셀에 스타일 적용
        for row in range(2, data_rows + 2):
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = data_border
        
        # 컬럼 너비 조정
        column_widths = {
            'A': 12,  # 일자
            'B': 8,   # 순번
            'C': 12,  # 거래처코드
            'D': 15,  # 거래처명
            'E': 10,  # 출하창고
            'F': 10,  # 담당자
            'G': 12,  # 품목코드
            'H': 20,  # 품목명
            'I': 15,  # 규격
            'J': 10,  # 수량
            'K': 12,  # 사용유형
            'L': 20   # 적요
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    def create_summary_sheet(self, wb, aggregated_data: Dict[str, Any], brand_name: str):
        """요약 시트 생성 (브랜드별)"""
        summary_ws = wb.create_sheet("요약")
        
        # 브랜드별 요약 정보
        aggregated_by_brand = aggregated_data.get("aggregated_by_brand", {})
        products = aggregated_by_brand.get(brand_name, [])
        
        if products:
            # 신뢰도를 정수로 변환하여 평균 계산
            confidences = []
            for p in products:
                conf = p.get("신뢰도", 0)
                if isinstance(conf, str):
                    try:
                        conf = float(conf)
                    except ValueError:
                        conf = 0
                confidences.append(conf)
            avg_confidence = sum(confidences) / len(confidences) if confidences else 0
            total_quantity = sum(int(p["총_수량"]) if isinstance(p["총_수량"], (int, str)) else 0 for p in products)
        else:
            avg_confidence = 0
            total_quantity = 0
        
        summary_data = [
            ["항목", "값"],
            ["브랜드", brand_name],
            ["총 제품 종류", len(products)],
            ["총 수량", total_quantity],
            ["평균 신뢰도", f"{avg_confidence:.1f}%"],
            ["처리된 스레드", len(aggregated_data.get("thread_summaries", []))],
            ["생성일시", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        # 요약 데이터 입력
        for i, row_data in enumerate(summary_data, start=1):
            for j, value in enumerate(row_data, start=1):
                summary_ws.cell(row=i, column=j, value=value)
        
        # 요약 시트 스타일
        for i in range(1, len(summary_data) + 1):
            for j in range(1, 3):
                cell = summary_ws.cell(row=i, column=j)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if i == 1:  # 헤더
                    cell.font = Font(bold=True)
        
        summary_ws.column_dimensions['A'].width = 15
        summary_ws.column_dimensions['B'].width = 20
    
    def validate_data(self, aggregated_data: Dict[str, Any]) -> Dict[str, Any]:
        """데이터 검증"""
        products = aggregated_data.get("aggregated_products", [])
        
        if not products:
            return {
                "average_confidence": 0,
                "thread_count": 0,
                "is_valid": False
            }
        
        avg_confidence = sum(p.get("신뢰도", 0) for p in products) / len(products)
        thread_count = len(aggregated_data.get("thread_summaries", []))
        
        return {
            "average_confidence": round(avg_confidence, 2),
            "thread_count": thread_count,
            "is_valid": avg_confidence > 50 and len(products) > 0
        }
    
    def generate_excel_with_summary(self, aggregated_data: Dict[str, Any], output_path: str) -> bool:
        """
        요약 시트가 포함된 Excel 파일 생성
        """
        try:
            # 새 워크북 생성
            wb = openpyxl.Workbook()
            
            # 메인 데이터 시트
            ws = wb.active
            ws.title = "주문서입력"
            
            # 헤더 및 데이터 설정
            self.setup_headers(ws)
            products = aggregated_data.get("aggregated_products", [])
            self.fill_data(ws, products)
            self.apply_styles(ws, len(products))
            
            # 요약 시트 추가
            self.create_summary_sheet(wb, aggregated_data)
            
            # 파일 저장
            wb.save(output_path)
            print(f"Excel 파일 생성 완료 (요약 포함): {output_path}")
            return True
            
        except Exception as e:
            print(f"Excel 생성 오류: {e}")
            return False

if __name__ == "__main__":
    # 테스트 실행
    generator = ExcelGenerator()
    
    # 테스트용 데이터
    test_data = {
        "aggregated_products": [
            {
                "품목코드": "100002",
                "제품명": "바루랩 10-히알루론산 블루 아쿠아 젤 크림 80ml",
                "총_수량": 10,
                "신뢰도": 85,
                "적요": "출고 처리"
            },
            {
                "품목코드": "100003", 
                "제품명": "바루랩 10-히알루론산 블루 아쿠아 클렌징 젤 200ml",
                "총_수량": 5,
                "신뢰도": 90,
                "적요": "출고 처리"
            }
        ],
        "thread_summaries": [
            {"thread_index": 0, "summary": "출고 처리", "product_count": 2}
        ]
    }
    
    success = generator.generate_excel_with_summary(test_data, "test_output.xlsx")
    print(f"테스트 결과: {'성공' if success else '실패'}")
